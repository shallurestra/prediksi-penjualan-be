from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Query, Path
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import List, Optional
import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import psycopg2, psycopg2.extras
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import datetime, timedelta
from dotenv import load_dotenv
import os, io, csv, math, json, re

load_dotenv()

# ── CONFIG ──────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     int(os.getenv("DB_PORT", 5432)),
    "dbname":   os.getenv("DB_NAME", "postgres"),
    "user":     os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", ""),
}
SECRET_KEY  = os.getenv("SECRET_KEY", "sawangan-secret-key-2025")
ALGORITHM   = os.getenv("ALGORITHM", "HS256")
TOKEN_EXP   = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 1440))

pwd_context   = CryptContext(schemes=["bcrypt"], deprecated="auto")
bearer_scheme = HTTPBearer(auto_error=False)

# ── INIT DB ──────────────────────────────────────────────────────────────────
def init_db():
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    nama VARCHAR(100) NOT NULL,
                    username VARCHAR(50) UNIQUE NOT NULL,
                    password VARCHAR(255) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS penjualan (
                    id SERIAL PRIMARY KEY,
                    tanggal DATE NOT NULL,
                    hari VARCHAR(20) NOT NULL,
                    nama_produk VARCHAR(100) NOT NULL,
                    stok INTEGER NOT NULL,
                    terjual INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            # Alter table to add user_id column if not exists
            cur.execute("""
                ALTER TABLE penjualan ADD COLUMN IF NOT EXISTS user_id INTEGER REFERENCES users(id) ON DELETE CASCADE;
            """)
            # ── NEW: datasets table to track each uploaded file ──
            cur.execute("""
                CREATE TABLE IF NOT EXISTS datasets (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                    table_name VARCHAR(150) UNIQUE NOT NULL,
                    original_filename VARCHAR(255) NOT NULL,
                    row_count INTEGER DEFAULT 0,
                    error_count INTEGER DEFAULT 0,
                    columns_info TEXT DEFAULT '',
                    date_range_start DATE,
                    date_range_end DATE,
                    unique_products INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            # ── NEW: user activity log ──
            cur.execute("""
                CREATE TABLE IF NOT EXISTS user_activity_log (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                    activity_type VARCHAR(50) NOT NULL,
                    description TEXT,
                    metadata JSONB DEFAULT '{}',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
        conn.commit()
    finally:
        conn.close()

# ── APP ──────────────────────────────────────────────────────────────────────
app = FastAPI(title="Analisis Penjualan Getuk API")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"],
    allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

init_db()

# ── DB HELPERS ───────────────────────────────────────────────────────────────
def db_conn():
    return psycopg2.connect(**DB_CONFIG)

def db_fetchall(q, p=None):
    conn = db_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(q, p); return cur.fetchall()
    finally: conn.close()

def db_fetchone(q, p=None):
    conn = db_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(q, p); return cur.fetchone()
    finally: conn.close()

def db_execute(q, p=None):
    conn = db_conn()
    try:
        with conn.cursor() as cur: cur.execute(q, p)
        conn.commit()
    finally: conn.close()

def db_execute_returning(q, p=None):
    """Execute a query and return the first row (useful for INSERT ... RETURNING)."""
    conn = db_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(q, p)
            result = cur.fetchone()
        conn.commit()
        return result
    finally: conn.close()

def db_executemany(q, rows):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            psycopg2.extras.execute_batch(cur, q, rows, page_size=500)
        conn.commit()
    finally: conn.close()

# ── ACTIVITY LOG HELPER ─────────────────────────────────────────────────────
def log_activity(user_id: int, activity_type: str, description: str, metadata: dict = None):
    """Log a user activity to the database."""
    meta_json = json.dumps(metadata or {})
    db_execute(
        "INSERT INTO user_activity_log (user_id, activity_type, description, metadata) VALUES (%s, %s, %s, %s::jsonb)",
        (user_id, activity_type, description, meta_json)
    )

# ── AUTH HELPERS ─────────────────────────────────────────────────────────────
def hash_password(pw): return pwd_context.hash(pw)
def verify_password(plain, hashed): return pwd_context.verify(plain, hashed)

def create_token(data: dict):
    payload = {**data, "exp": datetime.utcnow() + timedelta(minutes=TOKEN_EXP)}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(creds: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    if not creds:
        raise HTTPException(401, "Token tidak ditemukan")
    try:
        payload = jwt.decode(creds.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username: raise HTTPException(401, "Token tidak valid")
        user = db_fetchone("SELECT id,nama,username FROM users WHERE username=%s", (username,))
        if not user: raise HTTPException(401, "User tidak ditemukan")
        return dict(user)
    except JWTError:
        raise HTTPException(401, "Token tidak valid atau kadaluarsa")

# ── PYDANTIC MODELS ──────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str; password: str

class RegisterRequest(BaseModel):
    nama: str; username: str; password: str

class PenjualanRow(BaseModel):
    tanggal: str; hari: str; nama_produk: str; stok: int; terjual: int

# ── ANALYSIS HELPERS ─────────────────────────────────────────────────────────
DAYS_ORDER = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]

def infer_category(name: str) -> str:
    if not name: return "Lainnya"
    first = name.upper().split()[0] if name.split() else ""
    MAP = {"GETUK":"Getuk","MENDOAN":"Mendoan","DAGE":"Dage","GHOSTING":"Ghosting",
           "JAHAT":"Jahat","JUMBO":"Jumbo","MEDIUM":"Medium","MEGA":"Mega",
           "MOZARELLA":"Mozarella","RAKSASA":"Raksasa","SACU":"Sacu",
           "SMALL":"Small","SUPER":"Super Jumbo","FREE":"Free Man"}
    return MAP.get(first, first.title() or "Lainnya")

def parse_date(raw: str) -> str:
    raw = str(raw).strip().strip('"').strip("'")
    parts = raw.replace("-","/").split("/")
    if len(parts)==3:
        if len(parts[2])==4: return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        if len(parts[0])==4: return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return raw

DAYS_MAP = {"senin":"Senin","selasa":"Selasa","rabu":"Rabu","kamis":"Kamis",
            "jumat":"Jumat","sabtu":"Sabtu","minggu":"Minggu"}

def normalize_day(raw: str) -> str:
    return DAYS_MAP.get(raw.strip().lower(), raw.strip().title())

def build_full_data(source_table: str = "penjualan", user_id: int = None):
    """Build full data from a given table (default: penjualan)."""
    if source_table == "penjualan" and user_id is not None:
        rows = db_fetchall("SELECT tanggal,hari,nama_produk,stok,terjual FROM penjualan WHERE user_id = %s ORDER BY tanggal", (user_id,))
    else:
        rows = db_fetchall(f"SELECT tanggal,hari,nama_produk,stok,terjual FROM {source_table} ORDER BY tanggal")
    if not rows: return None
    tx, daily_map = [], {}
    for r in rows:
        tgl = r["tanggal"].strftime("%Y-%m-%d") if hasattr(r["tanggal"],"strftime") else str(r["tanggal"])
        hari, nama, stok, terjual = r["hari"], r["nama_produk"], int(r["stok"]), int(r["terjual"])
        tx.append({"nama":nama,"kategoriProduk":infer_category(nama),"stok":stok,"terjual":terjual,"hari":hari,"tanggal":tgl})
        key = f"{tgl}_{hari}"
        if key not in daily_map:
            daily_map[key] = {"tanggal":tgl,"hari":hari,"Total_Stok":0,"Total_Terjual":0,"Jumlah_Produk":0}
        daily_map[key]["Total_Stok"] += stok
        daily_map[key]["Total_Terjual"] += terjual
        daily_map[key]["Jumlah_Produk"] += 1
    daily = sorted(daily_map.values(), key=lambda x: x["tanggal"])
    return {
        "preprocessSummary": {
            "total": len(tx), "valid": len(tx), "invalid": 0,
            "startDate": daily[0]["tanggal"] if daily else "",
            "endDate":   daily[-1]["tanggal"] if daily else "",
            "uniqueProducts":   len(set(r["nama"] for r in tx)),
            "uniqueCategories": len(set(r["kategoriProduk"] for r in tx)),
            "categorySource": "Inferensi dari nama produk",
        },
        "transactionRows": tx,
        "dailyAggregated": daily,
    }

def run_kmeans_analysis(daily: list):
    """Jalankan K-Means sklearn pada data harian, kembalikan clusteredData + stats + elbowData."""
    if not daily: return None
    df = pd.DataFrame(daily)
    features = ["Total_Stok","Total_Terjual"]
    scaler = StandardScaler()
    X = scaler.fit_transform(df[features])

    # Elbow
    elbow = []
    for k in range(1, min(11, len(df)+1)):
        km = KMeans(n_clusters=k, random_state=42, n_init=10)
        km.fit(X); elbow.append({"k": k, "inertia": float(km.inertia_)})

    # Final k=3
    k_final = min(3, len(df))
    km = KMeans(n_clusters=k_final, random_state=42, n_init=10)
    df["_cluster"] = km.fit_predict(X)

    # Label by avg terjual
    means = df.groupby("_cluster")["Total_Terjual"].mean().sort_values()
    labels = {1:["Sedang"], 2:["Rendah","Tinggi"], 3:["Rendah","Sedang","Tinggi"]}
    label_list = labels.get(k_final, [f"Cluster {i}" for i in range(k_final)])
    label_map = {cid: label_list[i] for i, cid in enumerate(means.index)}
    df["cluster"] = df["_cluster"].map(label_map)

    clustered = df.drop(columns=["_cluster"]).to_dict(orient="records")

    stats = []
    for label in label_list:
        sub = df[df["cluster"]==label]
        if sub.empty: continue
        day_counts = sub["hari"].value_counts()
        stats.append({
            "label": label,
            "count": len(sub),
            "avgTerjual": float(sub["Total_Terjual"].mean()),
            "avgStok":    float(sub["Total_Stok"].mean()),
            "minTerjual": int(sub["Total_Terjual"].min()),
            "maxTerjual": int(sub["Total_Terjual"].max()),
            "dominantDay": day_counts.index[0] if len(day_counts) else "-",
        })
    return {"clusteredData": clustered, "stats": stats, "elbowData": elbow}

def build_forecast(clustered_data: list, tx_rows: list) -> list:
    """Buat 7-hari forecast dari clusteredData."""
    if not clustered_data: return []
    
    # Parse tanggal dengan berbagai format yang mungkin
    last_date_str = clustered_data[-1]["tanggal"]
    try:
        # Coba format YYYY-MM-DD
        last_date = datetime.strptime(last_date_str, "%Y-%m-%d")
    except ValueError:
        try:
            # Coba format MM/DD/YYYY
            last_date = datetime.strptime(last_date_str, "%m/%d/%Y")
        except ValueError:
            try:
                # Coba format DD/MM/YYYY
                last_date = datetime.strptime(last_date_str, "%d/%m/%Y")
            except ValueError:
                # Fallback ke hari ini
                last_date = datetime.now()

    all_counts = {"Rendah":0,"Sedang":0,"Tinggi":0}
    for d in clustered_data:
        if d["cluster"] in all_counts: all_counts[d["cluster"]] += 1
    overall = max(all_counts, key=all_counts.get)

    forecasts = []
    for i in range(1, 8):
        nd = last_date + timedelta(days=i)
        wd = nd.weekday()
        hari_map = {0:"Senin",1:"Selasa",2:"Rabu",3:"Kamis",4:"Jumat",5:"Sabtu",6:"Minggu"}
        hari = hari_map[wd]

        day_rows = [d for d in clustered_data if d["hari"]==hari]
        counts = {"Rendah":0,"Sedang":0,"Tinggi":0}
        for d in day_rows:
            if d["cluster"] in counts: counts[d["cluster"]] += 1

        sample = len(day_rows) or len(clustered_data)
        dom_cat = max(counts, key=counts.get) if day_rows else overall
        dom_rows = [d for d in day_rows if d["cluster"]==dom_cat] if day_rows else \
                   [d for d in clustered_data if d["cluster"]==overall]
        dom_count = counts.get(dom_cat, 0) if day_rows else len(dom_rows)
        dom_ratio = dom_count / sample if sample else 0

        avg_sold = sum(d["Total_Terjual"] for d in dom_rows)/len(dom_rows) if dom_rows else 0
        avg_stok = sum(d["Total_Stok"]    for d in dom_rows)/len(dom_rows) if dom_rows else 0
        min_sold = min((d["Total_Terjual"] for d in dom_rows), default=0)
        max_sold = max((d["Total_Terjual"] for d in dom_rows), default=0)
        min_stok = min((d["Total_Stok"]    for d in dom_rows), default=0)
        max_stok = max((d["Total_Stok"]    for d in dom_rows), default=0)

        conf = "Tinggi" if sample>=8 and dom_ratio>=0.6 else "Sedang" if sample>=4 and dom_ratio>=0.45 else "Rendah"
        basis = (f"Berdasarkan {sample} data historis hari {hari}, kategori dominan {dom_cat} "
                 f"muncul {dom_count} kali ({dom_ratio*100:.1f}% dari riwayat hari yang sama)."
                 if day_rows else "Menggunakan rata-rata keseluruhan data historis.")

        # Product recommendations
        sup = [t for t in tx_rows if t["hari"]==hari and t.get("cluster")==dom_cat]
        if not sup: sup = [t for t in tx_rows if t["hari"]==hari]
        if not sup: sup = tx_rows
        prod_map = {}
        for t in sup:
            k = t["nama"]
            if k not in prod_map: prod_map[k] = {"nama":k,"kategoriProduk":t["kategoriProduk"],"totalTerjual":0,"totalStok":0,"count":0,"days":set()}
            prod_map[k]["totalTerjual"] += t["terjual"]; prod_map[k]["totalStok"] += t["stok"]
            prod_map[k]["count"] += 1; prod_map[k]["days"].add(t["tanggal"])
        top_prods = sorted(prod_map.values(), key=lambda x: x["totalTerjual"], reverse=True)[:5]
        reko = [{"kategoriProduk":p["kategoriProduk"],"nama":p["nama"],
                 "rataRataTerjualProduk":round(p["totalTerjual"]/p["count"]) if p["count"] else 0,
                 "rekomendasiStokProduk":round(p["totalStok"]/p["count"]) if p["count"] else 0,
                 "jumlahHariAktif":len(p["days"])} for p in top_prods]
        cat_dom = top_prods[0]["kategoriProduk"] if top_prods else "Lainnya"

        forecasts.append({
            "tanggal": nd.strftime("%Y-%m-%d"), "hari": hari, "kategori": dom_cat,
            "perkiraanTerjual": round(avg_sold),
            "rekomendasiStok": round(avg_stok),
            "rentangPerkiraanPenjualan": f"{round(min_sold)} - {round(max_sold)}",
            "tingkatKepercayaan": conf,
            "dasarPerkiraan": basis,
            "kategoriProdukDominan": cat_dom,
            "produkUtamaPendukung": ", ".join(p["nama"] for p in top_prods[:3]) or "-",
            "rekomendasiProduk": reko,
            "jumlahRiwayatHariSerupa": sample,
            "jumlahRiwayatKategoriDominan": dom_count,
            "proporsiDominasiKategori": round(dom_ratio*100, 2),
            "rataRataTotalTerjualHistorisHariSama": round(sum(d["Total_Terjual"] for d in day_rows)/len(day_rows)) if day_rows else round(avg_sold),
            "rataRataTotalTerjualKategoriDominan": round(avg_sold),
            "minimumTotalTerjualHistoris": round(min_sold), "maksimumTotalTerjualHistoris": round(max_sold),
            "rataRataTotalStokHistorisHariSama": round(sum(d["Total_Stok"] for d in day_rows)/len(day_rows)) if day_rows else round(avg_stok),
            "rataRataTotalStokKategoriDominan": round(avg_stok),
            "minimumTotalStokHistoris": round(min_stok), "maksimumTotalStokHistoris": round(max_stok),
        })
    return forecasts

def build_forecast_for_period(clustered_data: list, tx_rows: list, start_date: datetime, end_date: datetime) -> list:
    """Buat forecast dari clusteredData untuk rentang tanggal tertentu."""
    if not clustered_data: return []
    
    all_counts = {"Rendah":0,"Sedang":0,"Tinggi":0}
    for d in clustered_data:
        if d["cluster"] in all_counts: all_counts[d["cluster"]] += 1
    overall = max(all_counts, key=all_counts.get)

    days_diff = (end_date - start_date).days + 1
    forecasts = []
    for i in range(days_diff):
        nd = start_date + timedelta(days=i)
        wd = nd.weekday()
        hari_map = {0:"Senin",1:"Selasa",2:"Rabu",3:"Kamis",4:"Jumat",5:"Sabtu",6:"Minggu"}
        hari = hari_map[wd]

        day_rows = [d for d in clustered_data if d["hari"]==hari]
        counts = {"Rendah":0,"Sedang":0,"Tinggi":0}
        for d in day_rows:
            if d["cluster"] in counts: counts[d["cluster"]] += 1

        sample = len(day_rows) or len(clustered_data)
        dom_cat = max(counts, key=counts.get) if day_rows else overall
        dom_rows = [d for d in day_rows if d["cluster"]==dom_cat] if day_rows else \
                   [d for d in clustered_data if d["cluster"]==overall]
        dom_count = counts.get(dom_cat, 0) if day_rows else len(dom_rows)
        dom_ratio = dom_count / sample if sample else 0

        avg_sold = sum(d["Total_Terjual"] for d in dom_rows)/len(dom_rows) if dom_rows else 0
        avg_stok = sum(d["Total_Stok"]    for d in dom_rows)/len(dom_rows) if dom_rows else 0
        min_sold = min((d["Total_Terjual"] for d in dom_rows), default=0)
        max_sold = max((d["Total_Terjual"] for d in dom_rows), default=0)
        min_stok = min((d["Total_Stok"]    for d in dom_rows), default=0)
        max_stok = max((d["Total_Stok"]    for d in dom_rows), default=0)

        conf = "Tinggi" if sample>=8 and dom_ratio>=0.6 else "Sedang" if sample>=4 and dom_ratio>=0.45 else "Rendah"
        basis = (f"Berdasarkan {sample} data historis hari {hari}, kategori dominan {dom_cat} "
                 f"muncul {dom_count} kali ({dom_ratio*100:.1f}% dari riwayat hari yang sama)."
                 if day_rows else "Menggunakan rata-rata keseluruhan data historis.")

        # Product recommendations
        sup = [t for t in tx_rows if t["hari"]==hari and t.get("cluster")==dom_cat]
        if not sup: sup = [t for t in tx_rows if t["hari"]==hari]
        if not sup: sup = tx_rows
        prod_map = {}
        for t in sup:
            k = t["nama"]
            if k not in prod_map: prod_map[k] = {"nama":k,"kategoriProduk":t["kategoriProduk"],"totalTerjual":0,"totalStok":0,"count":0,"days":set()}
            prod_map[k]["totalTerjual"] += t["terjual"]; prod_map[k]["totalStok"] += t["stok"]
            prod_map[k]["count"] += 1; prod_map[k]["days"].add(t["tanggal"])
        top_prods = sorted(prod_map.values(), key=lambda x: x["totalTerjual"], reverse=True)[:5]
        reko = [{"kategoriProduk":p["kategoriProduk"],"nama":p["nama"],
                 "rataRataTerjualProduk":round(p["totalTerjual"]/p["count"]) if p["count"] else 0,
                 "rekomendasiStokProduk":round(p["totalStok"]/p["count"]) if p["count"] else 0,
                 "jumlahHariAktif":len(p["days"])} for p in top_prods]
        cat_dom = top_prods[0]["kategoriProduk"] if top_prods else "Lainnya"

        forecasts.append({
            "tanggal": nd.strftime("%Y-%m-%d"), "hari": hari, "kategori": dom_cat,
            "perkiraanTerjual": round(avg_sold),
            "rekomendasiStok": round(avg_stok),
            "rentangPerkiraanPenjualan": f"{round(min_sold)} - {round(max_sold)}",
            "tingkatKepercayaan": conf,
            "dasarPerkiraan": basis,
            "kategoriProdukDominan": cat_dom,
            "produkUtamaPendukung": ", ".join(p["nama"] for p in top_prods[:3]) or "-",
            "rekomendasiProduk": reko,
            "jumlahRiwayatHariSerupa": sample,
            "jumlahRiwayatKategoriDominan": dom_count,
            "proporsiDominasiKategori": round(dom_ratio*100, 2),
            "rataRataTotalTerjualHistorisHariSama": round(sum(d["Total_Terjual"] for d in day_rows)/len(day_rows)) if day_rows else round(avg_sold),
            "rataRataTotalTerjualKategoriDominan": round(avg_sold),
            "minimumTotalTerjualHistoris": round(min_sold), "maksimumTotalTerjualHistoris": round(max_sold),
            "rataRataTotalStokHistorisHariSama": round(sum(d["Total_Stok"] for d in day_rows)/len(day_rows)) if day_rows else round(avg_stok),
            "rataRataTotalStokKategoriDominan": round(avg_stok),
            "minimumTotalStokHistoris": round(min_stok), "maksimumTotalStokHistoris": round(max_stok),
        })
    return forecasts

# ── DYNAMIC TABLE HELPERS ────────────────────────────────────────────────────
def sanitize_table_name(name: str) -> str:
    """Sanitize a table name to prevent SQL injection."""
    return re.sub(r'[^a-zA-Z0-9_]', '', name)

def create_dataset_table(table_name: str):
    """Create a new table for a specific dataset upload."""
    safe_name = sanitize_table_name(table_name)
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS {safe_name} (
                    id SERIAL PRIMARY KEY,
                    tanggal DATE NOT NULL,
                    hari VARCHAR(20) NOT NULL,
                    nama_produk VARCHAR(100) NOT NULL,
                    stok INTEGER NOT NULL,
                    terjual INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
        conn.commit()
    finally:
        conn.close()

def insert_into_dataset_table(table_name: str, rows: list):
    """Insert rows into a dataset-specific table."""
    safe_name = sanitize_table_name(table_name)
    db_executemany(
        f"INSERT INTO {safe_name} (tanggal,hari,nama_produk,stok,terjual) VALUES (%s,%s,%s,%s,%s)",
        rows
    )

# ── AUTH ENDPOINTS ───────────────────────────────────────────────────────────
@app.get("/")
def root(): return {"message": "API Analisis Penjualan Getuk berjalan!"}

@app.post("/api/auth/login")
def login(body: LoginRequest):
    user = db_fetchone("SELECT id,nama,username,password FROM users WHERE username=%s", (body.username,))
    if not user or not verify_password(body.password, user["password"]):
        raise HTTPException(401, "Username atau password salah")
    # Log login activity
    log_activity(user["id"], "login", f"User {user['username']} berhasil login")
    return {"access_token": create_token({"sub": user["username"]}), "token_type": "bearer",
            "user": {"id": user["id"], "nama": user["nama"], "username": user["username"]}}

@app.post("/api/auth/register")
def register(body: RegisterRequest):
    if len(body.username.strip()) < 4: raise HTTPException(400, "Username minimal 4 karakter")
    if len(body.password) < 6:         raise HTTPException(400, "Password minimal 6 karakter")
    if not body.nama.strip():           raise HTTPException(400, "Nama tidak boleh kosong")
    if db_fetchone("SELECT id FROM users WHERE username=%s", (body.username.strip(),)):
        raise HTTPException(400, "Username sudah digunakan")
    db_execute("INSERT INTO users (nama,username,password) VALUES (%s,%s,%s)",
               (body.nama.strip(), body.username.strip(), hash_password(body.password)))
    return {"message": "Akun berhasil dibuat"}

@app.get("/api/auth/me")
def get_me(u=Depends(get_current_user)): return u

# ── DATA ENDPOINTS ───────────────────────────────────────────────────────────
@app.get("/api/full-data")
def get_full_data(u=Depends(get_current_user)):
    data = build_full_data(user_id=u["id"])
    if not data: raise HTTPException(404, "Tidak ada data penjualan di database")
    return data

@app.get("/api/dataset-info")
def get_dataset_info(u=Depends(get_current_user)):
    row = db_fetchone("""
        SELECT COUNT(*) AS total_baris, COUNT(DISTINCT nama_produk) AS total_produk_unik,
               COUNT(DISTINCT tanggal) AS total_hari, MIN(tanggal)::text AS tanggal_mulai,
               MAX(tanggal)::text AS tanggal_akhir, SUM(terjual) AS total_terjual, SUM(stok) AS total_stok
        FROM penjualan WHERE user_id = %s""", (u["id"],))
    # If no data found, return defaults
    if not row or row.get("total_baris") == 0:
        return {"total_baris": 0, "total_produk_unik": 0, "total_hari": 0, "tanggal_mulai": None, "tanggal_akhir": None, "total_terjual": 0, "total_stok": 0}
    return dict(row)

@app.get("/api/daily-sales")
def get_daily_sales(u=Depends(get_current_user)):
    rows = db_fetchall("""
        SELECT tanggal::text AS "Tanggal", hari AS "Hari",
               SUM(terjual) AS "Total_Terjual", SUM(stok) AS "Total_Stok"
        FROM penjualan WHERE user_id = %s GROUP BY tanggal,hari ORDER BY tanggal""", (u["id"],))
    return [dict(r) for r in rows]

@app.get("/api/top-products")
def get_top_products(limit: int = 10, u=Depends(get_current_user)):
    rows = db_fetchall("""
        SELECT nama_produk AS "Nama Produk", SUM(terjual) AS "Total_Terjual",
               ROUND(AVG(terjual)::numeric,2) AS "Rata_Terjual"
        FROM penjualan WHERE user_id = %s GROUP BY nama_produk ORDER BY "Total_Terjual" DESC LIMIT %s""", (u["id"], limit))
    return [dict(r) for r in rows]

@app.get("/api/products")
def get_products(u=Depends(get_current_user)):
    rows = db_fetchall("""
        SELECT nama_produk AS "Nama Produk", SUM(stok) AS "Total_Stok",
               SUM(terjual) AS "Total_Terjual", ROUND(AVG(terjual)::numeric,2) AS "Rata_Terjual",
               COUNT(*) AS "Jumlah_Hari",
               ROUND(SUM(terjual)::numeric/NULLIF(SUM(stok),0),4) AS "Rasio_Terjual"
        FROM penjualan WHERE user_id = %s GROUP BY nama_produk ORDER BY "Total_Terjual" DESC""", (u["id"],))
    return {"total_produk": len(rows), "data": [dict(r) for r in rows]}

# ── ANALYZE ENDPOINT ─────────────────────────────────────────────────────────
@app.get("/api/analyze")
def analyze(u=Depends(get_current_user)):
    """Jalankan K-Means sklearn server-side, kembalikan clusteredData + stats + elbowData."""
    data = build_full_data(user_id=u["id"])
    if not data: raise HTTPException(404, "Tidak ada data")
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result: raise HTTPException(500, "Analisis gagal")
    # Enrich tx rows dengan cluster info
    cluster_lookup = {f"{d['tanggal']}_{d['hari']}": d["cluster"] for d in result["clusteredData"]}
    tx_enriched = [{**t, "cluster": cluster_lookup.get(f"{t['tanggal']}_{t['hari']}")} for t in data["transactionRows"]]
    return {**result, "transactionRows": tx_enriched, "preprocessSummary": data["preprocessSummary"]}

# ── FORECAST ENDPOINT ────────────────────────────────────────────────────────
@app.get("/api/forecast")
def forecast(u=Depends(get_current_user)):
    """Buat 7-hari forecast berdasarkan K-Means clustering."""
    data = build_full_data(user_id=u["id"])
    if not data: raise HTTPException(404, "Tidak ada data")
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result: raise HTTPException(500, "Analisis gagal")
    cluster_lookup = {f"{d['tanggal']}_{d['hari']}": d["cluster"] for d in result["clusteredData"]}
    tx_enriched = [{**t, "cluster": cluster_lookup.get(f"{t['tanggal']}_{t['hari']}")} for t in data["transactionRows"]]
    forecasts = build_forecast(result["clusteredData"], tx_enriched)
    return {"forecasts": forecasts, "total": len(forecasts)}

@app.get("/api/forecast/period")
def forecast_period(start_date: str = Query(...), end_date: str = Query(...), u=Depends(get_current_user)):
    """Buat forecast untuk periode tanggal kustom menggunakan data penjualan utama."""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Format tanggal harus YYYY-MM-DD")
    
    if start_dt > end_dt:
        raise HTTPException(400, "Tanggal mulai tidak boleh setelah tanggal akhir")
    
    days_diff = (end_dt - start_dt).days + 1
    if days_diff > 31:
        raise HTTPException(400, "Maksimum durasi prediksi adalah 31 hari")
        
    data = build_full_data(user_id=u["id"])
    if not data: raise HTTPException(404, "Tidak ada data penjualan")
    
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result: raise HTTPException(500, "Analisis gagal")
    
    cluster_lookup = {f"{d['tanggal']}_{d['hari']}": d["cluster"] for d in result["clusteredData"]}
    tx_enriched = [{**t, "cluster": cluster_lookup.get(f"{t['tanggal']}_{t['hari']}")} for t in data["transactionRows"]]
    
    forecasts = build_forecast_for_period(result["clusteredData"], tx_enriched, start_dt, end_dt)
    return {"forecasts": forecasts, "total": len(forecasts)}

# ── UPLOAD ENDPOINT ──────────────────────────────────────────────────────────
@app.post("/api/penjualan/upload")
async def upload_penjualan(file: UploadFile = File(...), u=Depends(get_current_user)):
    """Upload file CSV/Excel, parse, simpan ke DB (tabel penjualan + tabel baru per upload)."""
    content = await file.read()
    fname = file.filename.lower()
    original_filename = file.filename
    rows_parsed = []
    errors = 0

    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            raw_rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    except Exception as e:
        raise HTTPException(400, f"Gagal membaca file: {e}")

    def get_val(row, keys):
        for k in keys:
            for rk in row:
                if str(rk).strip().lower() == k.lower():
                    return row[rk]
        return None

    for row in raw_rows:
        try:
            nama   = str(get_val(row, ["Nama Produk","nama produk","produk"]) or "").strip()
            stok   = int(float(str(get_val(row, ["Stok","stok awal","stok"]) or 0)))
            terjual= int(float(str(get_val(row, ["Terjual","terjual"]) or 0)))
            hari   = normalize_day(str(get_val(row, ["Hari","hari","nama hari"]) or ""))
            tgl    = parse_date(str(get_val(row, ["Tanggal","tanggal","date"]) or ""))
            if nama and hari and tgl: rows_parsed.append((tgl, hari, nama, stok, terjual))
            else: errors += 1
        except: errors += 1

    if not rows_parsed:
        raise HTTPException(400, "Tidak ada baris valid dalam file. Pastikan kolom: Tanggal, Hari, Nama Produk, Stok, Terjual.")

    # ── 1. Insert ke tabel penjualan utama dengan user_id ──
    db_executemany(
        "INSERT INTO penjualan (tanggal,hari,nama_produk,stok,terjual,user_id) VALUES (%s,%s,%s,%s,%s,%s)",
        [(r[0], r[1], r[2], r[3], r[4], u["id"]) for r in rows_parsed]
    )

    # ── 2. Create new dynamic table for this upload ──
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    table_name = f"dataset_{u['id']}_{timestamp_str}"
    safe_table = sanitize_table_name(table_name)
    
    create_dataset_table(safe_table)
    insert_into_dataset_table(safe_table, rows_parsed)

    # ── 3. Compute metadata for the dataset ──
    dates_parsed = [r[0] for r in rows_parsed]
    products = set(r[2] for r in rows_parsed)
    
    # Register dataset
    dataset_record = db_execute_returning(
        """INSERT INTO datasets (user_id, table_name, original_filename, row_count, error_count, 
           columns_info, date_range_start, date_range_end, unique_products)
           VALUES (%s, %s, %s, %s, %s, %s, %s::date, %s::date, %s) RETURNING id""",
        (u["id"], safe_table, original_filename, len(rows_parsed), errors,
         "Tanggal, Hari, Nama Produk, Stok, Terjual",
         min(dates_parsed), max(dates_parsed), len(products))
    )
    dataset_id = dataset_record["id"] if dataset_record else None

    # ── 4. Log activity ──
    log_activity(u["id"], "upload", f"Upload file: {original_filename}", {
        "filename": original_filename,
        "dataset_id": dataset_id,
        "table_name": safe_table,
        "row_count": len(rows_parsed),
        "error_count": errors,
        "unique_products": len(products),
    })

    data = build_full_data(user_id=u["id"])
    return {
        "message": f"Berhasil import {len(rows_parsed)} baris, {errors} baris dilewati.",
        "imported": len(rows_parsed), "errors": errors,
        "dataset_id": dataset_id,
        **data
    }

@app.delete("/api/penjualan/clear")
def clear_penjualan(u=Depends(get_current_user)):
    """Hapus data penjualan milik user saat ini."""
    db_execute("DELETE FROM penjualan WHERE user_id = %s", (u["id"],))
    log_activity(u["id"], "clear_data", "Menghapus data penjualan akun ini")
    return {"message": "Semua data penjualan Anda berhasil dihapus"}

# ── ACTIVITY LOG ENDPOINTS ───────────────────────────────────────────────────
@app.get("/api/activity-log")
def get_activity_log(limit: int = 50, u=Depends(get_current_user)):
    """Ambil riwayat aktivitas user yang sedang login."""
    rows = db_fetchall(
        """SELECT al.id, al.activity_type, al.description, al.metadata, 
                  al.created_at, u.nama AS user_nama, u.username AS user_username
           FROM user_activity_log al
           JOIN users u ON al.user_id = u.id
           WHERE al.user_id = %s
           ORDER BY al.created_at DESC
           LIMIT %s""",
        (u["id"], limit)
    )
    result = []
    for r in rows:
        item = dict(r)
        # Format created_at for frontend
        if hasattr(item["created_at"], "isoformat"):
            item["created_at"] = item["created_at"].isoformat()
        # Ensure metadata is dict
        if isinstance(item["metadata"], str):
            try:
                item["metadata"] = json.loads(item["metadata"])
            except:
                item["metadata"] = {}
        result.append(item)
    return {"activities": result, "total": len(result)}

# ── DATASET ENDPOINTS ────────────────────────────────────────────────────────
@app.get("/api/datasets")
def get_datasets(u=Depends(get_current_user)):
    """Ambil daftar semua dataset yang pernah di-upload user."""
    rows = db_fetchall(
        """SELECT id, table_name, original_filename, row_count, error_count,
                  columns_info, date_range_start, date_range_end, unique_products, created_at
           FROM datasets
           WHERE user_id = %s
           ORDER BY created_at DESC""",
        (u["id"],)
    )
    result = []
    for r in rows:
        item = dict(r)
        if hasattr(item.get("created_at"), "isoformat"):
            item["created_at"] = item["created_at"].isoformat()
        if hasattr(item.get("date_range_start"), "isoformat"):
            item["date_range_start"] = item["date_range_start"].isoformat() if item["date_range_start"] else None
        if hasattr(item.get("date_range_end"), "isoformat"):
            item["date_range_end"] = item["date_range_end"].isoformat() if item["date_range_end"] else None
        result.append(item)
    return {"datasets": result, "total": len(result)}

@app.get("/api/datasets/{dataset_id}")
def get_dataset_detail(dataset_id: int = Path(...), u=Depends(get_current_user)):
    """Ambil detail sebuah dataset beserta datanya."""
    ds = db_fetchone(
        "SELECT * FROM datasets WHERE id=%s AND user_id=%s",
        (dataset_id, u["id"])
    )
    if not ds:
        raise HTTPException(404, "Dataset tidak ditemukan")
    
    ds = dict(ds)
    safe_table = sanitize_table_name(ds["table_name"])
    
    # Check if table exists
    exists = db_fetchone(
        "SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_name=%s)",
        (safe_table,)
    )
    if not exists or not exists.get("exists", False):
        raise HTTPException(404, "Tabel dataset tidak ditemukan di database")
    
    # Build full data from the dataset-specific table
    data = build_full_data(safe_table)
    
    # Format dates
    if hasattr(ds.get("created_at"), "isoformat"):
        ds["created_at"] = ds["created_at"].isoformat()
    if hasattr(ds.get("date_range_start"), "isoformat"):
        ds["date_range_start"] = ds["date_range_start"].isoformat() if ds["date_range_start"] else None
    if hasattr(ds.get("date_range_end"), "isoformat"):
        ds["date_range_end"] = ds["date_range_end"].isoformat() if ds["date_range_end"] else None
    
    return {
        "dataset": ds,
        **(data or {"transactionRows": [], "dailyAggregated": [], "preprocessSummary": None}),
    }

@app.get("/api/datasets/{dataset_id}/analyze")
def analyze_dataset(dataset_id: int = Path(...), u=Depends(get_current_user)):
    """Jalankan K-Means pada dataset spesifik."""
    ds = db_fetchone(
        "SELECT * FROM datasets WHERE id=%s AND user_id=%s",
        (dataset_id, u["id"])
    )
    if not ds:
        raise HTTPException(404, "Dataset tidak ditemukan")
    
    safe_table = sanitize_table_name(ds["table_name"])
    data = build_full_data(safe_table)
    if not data:
        raise HTTPException(404, "Dataset kosong")
    
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result:
        raise HTTPException(500, "Analisis gagal")
    
    # Enrich tx rows
    cluster_lookup = {f"{d['tanggal']}_{d['hari']}": d["cluster"] for d in result["clusteredData"]}
    tx_enriched = [{**t, "cluster": cluster_lookup.get(f"{t['tanggal']}_{t['hari']}")} for t in data["transactionRows"]]
    
    # Build forecast
    forecasts = build_forecast(result["clusteredData"], tx_enriched)
    
    # Log analyze activity
    log_activity(u["id"], "analyze", f"Analisis dataset: {ds['original_filename']}", {
        "dataset_id": dataset_id,
        "filename": ds["original_filename"],
    })
    
    return {
        **result,
        "transactionRows": tx_enriched,
        "preprocessSummary": data["preprocessSummary"],
        "forecasts": forecasts,
    }

@app.get("/api/datasets/{dataset_id}/forecast/period")
def dataset_forecast_period(dataset_id: int = Path(...), start_date: str = Query(...), end_date: str = Query(...), u=Depends(get_current_user)):
    """Buat forecast untuk periode kustom berdasarkan data dari dataset spesifik."""
    ds = db_fetchone(
        "SELECT * FROM datasets WHERE id=%s AND user_id=%s",
        (dataset_id, u["id"])
    )
    if not ds:
        raise HTTPException(404, "Dataset tidak ditemukan")
    
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Format tanggal harus YYYY-MM-DD")
    
    if start_dt > end_dt:
        raise HTTPException(400, "Tanggal mulai tidak boleh setelah tanggal akhir")
    
    days_diff = (end_dt - start_dt).days + 1
    if days_diff > 31:
        raise HTTPException(400, "Maksimum durasi prediksi adalah 31 hari")
        
    safe_table = sanitize_table_name(ds["table_name"])
    data = build_full_data(safe_table)
    if not data:
        raise HTTPException(404, "Dataset kosong")
        
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result:
        raise HTTPException(500, "Analisis gagal")
        
    cluster_lookup = {f"{d['tanggal']}_{d['hari']}": d["cluster"] for d in result["clusteredData"]}
    tx_enriched = [{**t, "cluster": cluster_lookup.get(f"{t['tanggal']}_{t['hari']}")} for t in data["transactionRows"]]
    
    forecasts = build_forecast_for_period(result["clusteredData"], tx_enriched, start_dt, end_dt)
    return {"forecasts": forecasts, "total": len(forecasts)}

# ── EXPORT HELPERS ───────────────────────────────────────────────────────────
def make_excel_response(df_dict: dict, filename: str) -> StreamingResponse:
    """Buat StreamingResponse berisi file Excel dari dict of {sheet_name: list_of_dicts}."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, records in df_dict.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not records: ws.append(["Tidak ada data"]); continue
        headers = list(records[0].keys())
        ws.append(headers)
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C0392B")
            cell.alignment = Alignment(horizontal="center")
        for rec in records:
            ws.append([rec.get(h,"") for h in headers])
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# ── EXPORT ENDPOINTS ─────────────────────────────────────────────────────────
@app.get("/api/export/history")
def export_history(u=Depends(get_current_user)):
    """Export riwayat penjualan ke Excel (2 sheet: harian & per produk)."""
    rows = db_fetchall("SELECT tanggal::text,hari,nama_produk,stok,terjual FROM penjualan WHERE user_id = %s ORDER BY tanggal,nama_produk", (u["id"],))
    daily = db_fetchall("""
        SELECT tanggal::text AS "Tanggal", hari AS "Hari",
               SUM(stok) AS "Total Stok", SUM(terjual) AS "Total Terjual", COUNT(*) AS "Jumlah Produk"
        FROM penjualan WHERE user_id = %s GROUP BY tanggal,hari ORDER BY tanggal""", (u["id"],))
    produk = [{"Tanggal":r["tanggal"],"Hari":r["hari"],"Nama Produk":r["nama_produk"],
               "Stok":r["stok"],"Terjual":r["terjual"]} for r in rows]
    harian = [dict(r) for r in daily]
    log_activity(u["id"], "export", "Export history ke Excel")
    return make_excel_response({"Data Harian": harian, "Data Per Produk": produk},
                               f"history_penjualan_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.get("/api/export/report")
def export_report(u=Depends(get_current_user)):
    """Export laporan K-Means ke Excel."""
    data = build_full_data(user_id=u["id"])
    if not data: raise HTTPException(404, "Tidak ada data")
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result: raise HTTPException(500, "Analisis gagal")

    clustered = [{"Tanggal":d["tanggal"],"Hari":d["hari"],"Total Stok":d["Total_Stok"],
                  "Total Terjual":d["Total_Terjual"],"Jumlah Produk":d["Jumlah_Produk"],
                  "Kategori":d["cluster"]} for d in result["clusteredData"]]
    stats = [{"Kategori":s["label"],"Jumlah Hari":s["count"],
              "Rata-rata Terjual":round(s["avgTerjual"]),"Rata-rata Stok":round(s["avgStok"]),
              "Min Terjual":s["minTerjual"],"Max Terjual":s["maxTerjual"],"Hari Dominan":s["dominantDay"]}
             for s in result["stats"]]
    log_activity(u["id"], "export", "Export report K-Means ke Excel")
    return make_excel_response({"Ringkasan Cluster": stats, "Data Klaster Harian": clustered},
                               f"report_kmeans_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.get("/api/export/forecast")
def export_forecast(u=Depends(get_current_user)):
    """Export hasil forecast 7 hari ke Excel."""
    data = build_full_data(user_id=u["id"])
    if not data: raise HTTPException(404, "Tidak ada data")
    result = run_kmeans_analysis(data["dailyAggregated"])
    if not result: raise HTTPException(500, "Analisis gagal")
    cluster_lookup = {f"{d['tanggal']}_{d['hari']}": d["cluster"] for d in result["clusteredData"]}
    tx_enriched = [{**t, "cluster": cluster_lookup.get(f"{t['tanggal']}_{t['hari']}")} for t in data["transactionRows"]]
    forecasts = build_forecast(result["clusteredData"], tx_enriched)

    main_rows = [{"Tanggal":f["tanggal"],"Hari":f["hari"],"Kategori":f["kategori"],
                  "Perkiraan Terjual":f["perkiraanTerjual"],"Rekomendasi Stok":f["rekomendasiStok"],
                  "Rentang Terjual":f["rentangPerkiraanPenjualan"],"Tingkat Kepercayaan":f["tingkatKepercayaan"],
                  "Kategori Produk Dominan":f["kategoriProdukDominan"],
                  "Produk Utama":f["produkUtamaPendukung"],"Dasar Perkiraan":f["dasarPerkiraan"]}
                 for f in forecasts]
    detail_rows = []
    for f in forecasts:
        for p in f.get("rekomendasiProduk", []):
            detail_rows.append({"Tanggal":f["tanggal"],"Hari":f["hari"],"Nama Produk":p["nama"],
                                 "Kategori Produk":p["kategoriProduk"],
                                 "Rata-rata Terjual":p["rataRataTerjualProduk"],
                                 "Rekomendasi Stok":p["rekomendasiStokProduk"],
                                 "Hari Aktif":p["jumlahHariAktif"]})
    log_activity(u["id"], "export", "Export forecast ke Excel")
    return make_excel_response({"Forecast 7 Hari": main_rows, "Detail Produk": detail_rows},
                               f"forecast_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.get("/api/export/template")
def export_template(u=Depends(get_current_user)):
    """Download template Excel untuk upload data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Data Penjualan"
    headers = ["Tanggal","Hari","Nama Produk","Stok","Terjual"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = Alignment(horizontal="center")
    ws.append(["2025-10-01","Rabu","GETUK ORI 1 KG",100,75])
    for col, w in zip("ABCDE", [14,10,25,10,10]):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet("Panduan")
    guide = [["Kolom","Keterangan","Contoh"],
             ["Tanggal","Format YYYY-MM-DD atau D/M/YYYY","2025-10-01"],
             ["Hari","Nama hari Bahasa Indonesia","Senin, Selasa, Rabu, Kamis, Jumat, Sabtu, Minggu"],
             ["Nama Produk","Nama produk yang dijual","GETUK ORI 1 KG"],
             ["Stok","Jumlah stok awal (angka)","100"],
             ["Terjual","Jumlah terjual (angka)","75"]]
    for row in guide: ws2.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="template_penjualan.xlsx"'})
